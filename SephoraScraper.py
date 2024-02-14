import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time

from selenium.webdriver.common.proxy import Proxy, ProxyType

from webdrivermanager import ChromeDriverManager

from openpyxl import load_workbook

driver = webdriver.Chrome()

df = pd.read_excel('sephoralinks.xlsx')
product_links = df['Links'].tolist()

brand = []
name = []
price = []
size = []
num_ratings = []
rating = []
solutions = [] #ex. dryness

'''brand = df['Brand'].tolist()
name = df['ProductName'].tolist()
price = df['Price'].tolist()
size = df['Size'].tolist()
num_ratings = df['NumberOfRatings'].tolist()
rating = df['Rating'].tolist()
solutions = df['Solutions'].tolist()'''

wb = load_workbook("sephoralinks.xlsx")
ws = wb["Sheet1"]


for link in product_links[488:489]:
    row = str(product_links.index(link) + 2)

    driver.get(link)
    
    b = driver.find_element(By.CSS_SELECTOR, "a[data-at=brand_name]")

    ws['B'+row] = b.get_attribute("innerHTML")
    
    n = driver.find_element(By.CSS_SELECTOR, "span[data-at=product_name]")

    print(n.get_attribute("innerHTML"))
    ws['C'+row] = n.get_attribute("innerHTML")
    
    
    if driver.find_elements(By.CSS_SELECTOR, "b[class=css-0]"):
        pri = driver.find_element(By.CSS_SELECTOR, "b[class=css-0]")
        pr = pri.get_attribute("innerHTML")
    elif driver.find_elements(By.CSS_SELECTOR, "b[class=css-5fq4jh]"):
        pri = driver.find_element(By.CSS_SELECTOR, "b[class=css-5fq4jh]")
        pr = pri.get_attribute("innerHTML")
    else: pr = "None"

    ws['D'+row] = pr
    
    
    if driver.find_elements(By.CSS_SELECTOR, "span[class=css-15ro776]"):
        si = driver.find_element(By.CSS_SELECTOR, "span[class=css-15ro776]")
        #size.append(si.get_attribute("innerHTML"))
        ws['E'+row]= si.get_attribute("innerHTML")
    else: 
        si = "None"
        ws['E'+row]= si
    
    
    r = driver.find_element(By.CSS_SELECTOR, "span[class=css-1tbjoxk")
   
    ws['F'+row] = r.get_attribute("aria-label")
    
    
    num = driver.find_element(By.CSS_SELECTOR, "span[class=css-1j53ife]")
    
    ws['G'+row] = num.get_attribute("innerHTML")
    
    
    time.sleep(3)
    
    driver.execute_script("window.scrollTo(0, 1050)")
    
    time.sleep(3)
    
    driver.find_element(By.XPATH, "//button[@class='css-5fs8cb eanm77i0']").click()

    
    sol = driver.find_element(By.XPATH, "//div[@class='css-1w8ckn1 eanm77i0']")
    sol_text = str(sol.get_attribute("innerHTML"))
    if "Solutions for:" in sol_text:
        try : 
            #solutions.append(sol_text.split("Solutions for:")[1].split("</p")[0])
            ws['H'+row] = sol_text.split("Solutions for:")[1].split("</p")[0]
        except: 
            #solutions.append(sol_text.split("Solutions for:")[1].split("<b")[0])
            ws['H'+row] = sol_text.split("Solutions for:")[1].split("<b")[0]
    elif "Skincare Concerns:" in sol_text:
        try: 
            #solutions.append(sol_text.split("Skincare Concerns:")[1].split("<b")[0])
            ws['H'+row] = sol_text.split("Skincare Concerns:")[1].split("<b")[0]
        except: 
            #solutions.append(sol_text.split("Skincare Concerns:")[1].split("</p")[0])
            ws['H'+row] = sol_text.split("Skincare Concerns:")[1].split("</p")[0]
    else: 
        #solutions.append("None")
        ws['H'+row] = 'None'
    
    
    
    wb.save("sephoralinks.xlsx")
    
    
    
print(brand)
#print(len(brand))
print(name)
#print(len(name))
print(price)
#print(len(price))
print(size)
#print(len(size))
print(rating)
#print(len(rating))
print(num_ratings)
#print(len(num_ratings))
print(solutions)
#print(len(solutions))
    

print("done")

keyword = input("enter a character or press enter to continue")

driver.quit()
